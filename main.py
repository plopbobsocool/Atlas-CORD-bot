import discord
from discord.ext import commands
from openpyxl import Workbook, load_workbook
import os

intents = discord.Intents.default()
intents.messages = True
bot = commands.Bot(command_prefix='!', intents=intents)

FILE_NAME = 'Atlas Cords.xlsx'

def setup_excel():
    """Ensure the Excel file exists and is correctly formatted."""
    if not os.path.exists(FILE_NAME):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Messages"
        sheet.append(["Timestamp", "Username", "User ID", "Message"])
        workbook.save(FILE_NAME)

def append_to_excel(data):
    """Append a row of data to the Excel file."""
    try:
        workbook = load_workbook(FILE_NAME)
        sheet = workbook.active
        sheet.append(data)
        workbook.save(FILE_NAME)
    except Exception as e:
        print(f"Error writing to Excel file: {e}")

@bot.event
async def on_ready():
    """Triggers when the bot is ready."""
    print(f"Logged in as {bot.user}")
    setup_excel()

@bot.event
async def on_message(message):
    """Handles new messages."""
    if message.author.bot:
        return

    # Extract message details
    timestamp = message.created_at.strftime("%Y-%m-%d %H:%M:%S")
    username = message.author.name
    user_id = message.author.id
    content = message.content

    # Log the message to the Excel file
    append_to_excel([timestamp, username, user_id, content])
    print(f"Message logged: {content}")

TOKEN = "YOUR_BOT_TOKEN_HERE"
bot.run(TOKEN)
